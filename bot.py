import logging
import pandas as pd
import numpy as np
from scipy.optimize import minimize
from scipy.special import expit
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

try:
    from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
    from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, ConversationHandler, filters, ContextTypes
except ImportError as e:
    print("ERROR: Please install python-telegram-bot version 22.5")
    print("Run: pip install python-telegram-bot==22.5")
    raise e

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("telegram").setLevel(logging.WARNING)

BOT_TOKEN = "8578690075:AAFUHQQTPyM9tUlwtKxpV8go_YLPRGcMsog"

WAITING_CONFIG, WAITING_FILE = range(2)

DEFAULT_CONFIG = {
    'n_closed': 35,
    'n_open': 10,
    'open_max_score': 10,
    'passing_percentage': 60.0,
    'tier_thresholds': [93, 87, 80, 72, 67, 60],
    'tier_names': ['A+', 'A', 'B+', 'B', 'C+', 'C', 'D'],
    'tier_colors': ['#FF6B6B', '#FFA500', '#FFD700', '#90EE90', '#87CEEB', '#98FB98', '#D3D3D3'],
    'closed_weight': 1.0,
    'open_weight': 1.0,
    'use_percentage_grading': True,
}


class PartialCreditModel:
    """
    Partial Credit Model (PCM) Implementation for Item Response Theory
    Properly handles both dichotomous (0/1) and polytomous (0,1,2,...,m) responses
    """

    def __init__(self):
        self.item_difficulties = None
        self.step_difficulties = None
        self.person_abilities = None
        self.n_items = None
        self.n_persons = None
        self.max_scores = None
        self.convergence_success = False

    def safe_exp(self, x):
        """Numerically stable exponential with clipping"""
        x_clipped = np.clip(x, -700, 700)
        return np.exp(x_clipped)

    def pcm_probability(self, ability, difficulties, max_score):
        """
        Calculate probability of each score category using PCM
        For dichotomous items (max_score=1), this reduces to the Rasch model

        P(X=x|θ,δ) = exp(Σ(θ-δ_j)) / Σ exp(Σ(θ-δ_k)) for k=0 to m
        """
        if max_score == 1:
            logit = ability - difficulties[0]
            prob_1 = expit(logit)
            return np.array([1 - prob_1, prob_1])

        numerators = np.zeros(max_score + 1)
        numerators[0] = 1.0

        for x in range(1, max_score + 1):
            sum_diff = sum(ability - difficulties[j] for j in range(x))
            numerators[x] = self.safe_exp(sum_diff)

        denominator = np.sum(numerators)
        if denominator < 1e-300:
            denominator = 1e-300

        return numerators / denominator

    def log_likelihood(self, params, responses, max_scores, item_param_indices):
        """Calculate log-likelihood for Joint Maximum Likelihood Estimation"""
        n_persons, n_items = responses.shape
        abilities = params[:n_persons]

        ll = 0
        for i in range(n_persons):
            ability = np.clip(abilities[i], -10, 10)
            for j in range(n_items):
                response = responses[i, j]
                if np.isnan(response):
                    continue

                max_score = max_scores[j]
                start_idx, end_idx = item_param_indices[j]
                difficulties = params[n_persons + start_idx:n_persons + end_idx]

                probs = self.pcm_probability(ability, difficulties, max_score)

                response_int = int(round(response * max_score))
                response_int = np.clip(response_int, 0, max_score)

                prob = probs[response_int]
                ll += np.log(max(prob, 1e-15))

        return -ll

    def calibrate(self, responses, max_scores, max_iter=1000):
        """
        Calibrate PCM using Joint Maximum Likelihood Estimation
        responses: numpy array (n_persons x n_items) with values 0 to 1 (normalized)
        max_scores: array of maximum score for each item (1 for dichotomous, >1 for polytomous)
        """
        self.n_persons, self.n_items = responses.shape
        self.max_scores = np.array(max_scores)

        item_param_indices = []
        total_params = 0
        for j in range(self.n_items):
            n_steps = self.max_scores[j]
            item_param_indices.append((total_params, total_params + n_steps))
            total_params += n_steps

        person_scores = np.nanmean(responses, axis=1)
        init_abilities = np.log((person_scores + 0.01) / (1 - person_scores + 0.01))
        init_abilities = np.clip(init_abilities, -5, 5)

        init_difficulties = []
        for j in range(self.n_items):
            item_mean = np.nanmean(responses[:, j])
            base_diff = -np.log((item_mean + 0.01) / (1 - item_mean + 0.01))
            base_diff = np.clip(base_diff, -5, 5)

            n_steps = self.max_scores[j]
            for step in range(n_steps):
                step_adjustment = (step - n_steps/2) * 0.5
                init_difficulties.append(base_diff + step_adjustment)

        init_params = np.concatenate([init_abilities, init_difficulties])

        result = minimize(
            self.log_likelihood,
            init_params,
            args=(responses, self.max_scores, item_param_indices),
            method='L-BFGS-B',
            options={'maxiter': max_iter, 'disp': False}
        )

        self.convergence_success = result.success

        self.person_abilities = result.x[:self.n_persons]

        self.item_difficulties = []
        self.step_difficulties = []
        for j in range(self.n_items):
            start_idx, end_idx = item_param_indices[j]
            steps = result.x[self.n_persons + start_idx:self.n_persons + end_idx]
            self.step_difficulties.append(steps)
            self.item_difficulties.append(np.mean(steps))

        self.item_difficulties = np.array(self.item_difficulties)

        mean_difficulty = np.mean(self.item_difficulties)
        self.item_difficulties -= mean_difficulty
        self.person_abilities -= mean_difficulty
        for j in range(self.n_items):
            self.step_difficulties[j] -= mean_difficulty

        return self.person_abilities, self.item_difficulties, self.step_difficulties

    def estimate_ability(self, response_pattern, max_scores):
        """Estimate ability for a single person given calibrated item parameters"""
        def neg_likelihood(ability):
            ability = np.clip(ability[0], -10, 10)
            ll = 0
            for j, response in enumerate(response_pattern):
                if np.isnan(response):
                    continue

                max_score = max_scores[j]
                difficulties = self.step_difficulties[j]
                probs = self.pcm_probability(ability, difficulties, max_score)

                response_int = int(round(response * max_score))
                response_int = np.clip(response_int, 0, max_score)

                ll += np.log(max(probs[response_int], 1e-15))
            return -ll

        raw_score = np.nanmean(response_pattern)
        init_ability = np.log((raw_score + 0.01) / (1 - raw_score + 0.01))
        init_ability = np.clip(init_ability, -5, 5)

        result = minimize(neg_likelihood, [init_ability], method='L-BFGS-B',
                         bounds=[(-10, 10)])

        return result.x[0]


class AssessmentProcessor:
    """Process assessment data and generate reports with teacher configuration"""

    def __init__(self, config=None):
        self.pcm = PartialCreditModel()
        self.config = config or DEFAULT_CONFIG.copy()
        self.correct_answers = None
        self.total_questions = self.config['n_closed'] + self.config['n_open']
        self.has_fractional_warning = False

    def update_config(self, config):
        """Update configuration from teacher input"""
        self.config.update(config)
        self.total_questions = self.config['n_closed'] + self.config['n_open']

    def parse_excel(self, file_bytes):
        """Parse Excel file with correct answers and student responses"""
        df = pd.read_excel(io.BytesIO(file_bytes))

        self.correct_answers = df.iloc[0, 1:].values
        student_data = df.iloc[1:].copy()

        return student_data

    def score_responses(self, student_data):
        """
        Score student responses against correct answers
        Open question scores are rounded to nearest integer for proper PCM calibration
        """
        scored_data = []
        n_closed = self.config['n_closed']
        open_max = self.config['open_max_score']
        closed_weight = self.config['closed_weight']
        open_weight = self.config['open_weight']

        has_fractional_scores = False

        for idx, row in student_data.iterrows():
            student_id = row.iloc[0]
            responses = row.iloc[1:].values

            scored_responses = []
            raw_score = 0
            max_possible = 0

            for i, (student_ans, correct_ans) in enumerate(zip(responses, self.correct_answers)):
                if i < n_closed:
                    score = 1.0 if str(student_ans).strip().upper() == str(correct_ans).strip().upper() else 0.0
                    weighted_score = score * closed_weight
                    max_possible += closed_weight
                else:
                    try:
                        raw_open_score = float(student_ans)
                        raw_open_score = np.clip(raw_open_score, 0, open_max)

                        if raw_open_score != int(raw_open_score):
                            has_fractional_scores = True
                            raw_open_score = round(raw_open_score)

                        score = raw_open_score / open_max
                    except (ValueError, TypeError):
                        score = 0.0
                    weighted_score = score * open_weight
                    max_possible += open_weight

                scored_responses.append(score)
                raw_score += weighted_score

            scored_data.append({
                'student_id': student_id,
                'responses': scored_responses,
                'raw_score': raw_score,
                'max_possible': max_possible,
                'percentage': (raw_score / max_possible * 100) if max_possible > 0 else 0
            })

        self.has_fractional_warning = has_fractional_scores
        return scored_data

    def get_max_scores(self):
        """Get maximum score array for each item (1 for closed, configured for open)"""
        n_closed = self.config['n_closed']
        n_open = self.config['n_open']
        open_max = self.config['open_max_score']

        max_scores = [1] * n_closed + [open_max] * n_open
        return max_scores

    def create_tiers(self, abilities, percentages):
        """
        Create tier rankings based on PERCENTAGE SCORE (not percentile ranking)
        Students with the SAME percentage get the SAME tier

        Grading scale (based on score percentage):
        A+: ≥93%
        A:  87-93%
        B+: 80-87%
        B:  72-80%
        C+: 67-72%
        C:  60-67%
        D:  <60%
        """
        n_students = len(abilities)
        tiers = np.empty(n_students, dtype=object)

        tier_thresholds = self.config.get('tier_thresholds', [93, 87, 80, 72, 67, 60])
        tier_names = self.config['tier_names']

        for i in range(n_students):
            pct = percentages[i]

            tier_idx = len(tier_thresholds)
            for j, threshold in enumerate(tier_thresholds):
                if pct >= threshold:
                    tier_idx = j
                    break

            tiers[i] = tier_names[min(tier_idx, len(tier_names) - 1)]

        return tiers

    def calculate_ranks(self, percentages):
        """
        Calculate ranks where students with same percentage get same rank
        """
        n_students = len(percentages)
        rounded_percentages = np.round(percentages, 1)

        sorted_indices = np.argsort(rounded_percentages)[::-1]

        ranks = np.zeros(n_students, dtype=int)

        current_rank = 1
        i = 0
        while i < n_students:
            current_pct = rounded_percentages[sorted_indices[i]]

            same_pct_count = 1
            while i + same_pct_count < n_students and rounded_percentages[sorted_indices[i + same_pct_count]] == current_pct:
                same_pct_count += 1

            for j in range(same_pct_count):
                ranks[sorted_indices[i + j]] = current_rank

            i += same_pct_count
            current_rank += same_pct_count

        return ranks

    def generate_report(self, scored_data, abilities, difficulties, tiers):
        """Generate comprehensive Excel report"""
        percentages = np.array([d['percentage'] for d in scored_data])
        ranks = self.calculate_ranks(percentages)

        passing_pct = self.config['passing_percentage']

        results = []
        for i, data in enumerate(scored_data):
            pct = data['percentage']
            status = "PASS" if pct >= passing_pct else "FAIL"

            results.append({
                'Rank': ranks[i],
                'Student ID': data['student_id'],
                'Raw Score': f"{data['raw_score']:.2f}/{data['max_possible']:.1f}",
                'Percentage': f"{pct:.1f}%",
                'Ability (θ)': f"{abilities[i]:.3f}",
                'Tier': tiers[i],
                'Status': status
            })

        results_df = pd.DataFrame(results).sort_values('Rank')

        item_analysis = []
        max_scores = self.get_max_scores()
        n_closed = self.config['n_closed']

        for i in range(self.total_questions):
            q_type = "Closed" if i < n_closed else "Open"
            responses = [d['responses'][i] for d in scored_data]

            avg_score = np.mean(responses) * 100

            item_analysis.append({
                'Question': f"Q{i+1}",
                'Type': q_type,
                'Max Score': max_scores[i],
                'Difficulty (β)': f"{difficulties[i]:.3f}",
                'Avg Score %': f"{avg_score:.1f}%",
                'Discrimination': self._calculate_discrimination(responses, abilities)
            })

        item_df = pd.DataFrame(item_analysis)

        return results_df, item_df

    def _calculate_discrimination(self, item_responses, abilities):
        """Calculate item discrimination (point-biserial/Pearson correlation)"""
        if len(set(item_responses)) <= 1:
            return "N/A"
        correlation = np.corrcoef(item_responses, abilities)[0, 1]
        if np.isnan(correlation):
            return "N/A"
        return f"{correlation:.3f}"

    def create_excel_output(self, results_df, item_df, config_summary):
        """Create formatted Excel file with configuration summary"""
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            results_df.to_excel(writer, sheet_name='Student Results', index=False)
            item_df.to_excel(writer, sheet_name='Item Analysis', index=False)

            workbook = writer.book

            ws1 = workbook['Student Results']
            self._format_worksheet(ws1, results_df, highlight_status=True)

            ws2 = workbook['Item Analysis']
            self._format_worksheet(ws2, item_df)

            ws3 = workbook.create_sheet('Summary')
            self._create_summary_sheet(ws3, results_df, item_df, config_summary)

            ws4 = workbook.create_sheet('Configuration')
            self._create_config_sheet(ws4, config_summary)

        output.seek(0)
        return output

    def _format_worksheet(self, ws, df, highlight_status=False):
        """Apply formatting to worksheet with tier colors matching the grading scale"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        tier_fills = {
            'A+': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            'A': PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
            'B+': PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
            'B': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
            'C+': PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
            'C': PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid"),
            'D': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
        }

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        if highlight_status and 'Tier' in df.columns:
            tier_col = list(df.columns).index('Tier') + 1
            for row in range(2, len(df) + 2):
                cell = ws.cell(row=row, column=tier_col)
                tier_value = str(cell.value) if cell.value else ""
                if tier_value in tier_fills:
                    cell.fill = tier_fills[tier_value]
                    cell.font = Font(bold=True)

        if highlight_status and 'Status' in df.columns:
            status_col = list(df.columns).index('Status') + 1
            for row in range(2, len(df) + 2):
                cell = ws.cell(row=row, column=status_col)
                if cell.value == "PASS":
                    cell.fill = pass_fill
                elif cell.value == "FAIL":
                    cell.fill = fail_fill

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        ws.freeze_panes = 'A2'

    def _create_summary_sheet(self, ws, results_df, item_df, config_summary):
        """Create summary statistics sheet"""
        passing_pct = self.config['passing_percentage']
        pass_count = sum(1 for _, row in results_df.iterrows() if row['Status'] == 'PASS')
        fail_count = len(results_df) - pass_count
        pass_rate = (pass_count / len(results_df) * 100) if len(results_df) > 0 else 0

        summary_data = [
            ['ASSESSMENT SUMMARY REPORT', ''],
            ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['', ''],
            ['CONFIGURATION', ''],
            ['Closed Questions:', config_summary['n_closed']],
            ['Open Questions:', config_summary['n_open']],
            ['Open Question Max Score:', config_summary['open_max_score']],
            ['Passing Threshold:', f"{passing_pct}%"],
            ['', ''],
            ['STUDENT STATISTICS', ''],
            ['Total Students:', len(results_df)],
            ['Passed:', f"{pass_count} ({pass_rate:.1f}%)"],
            ['Failed:', f"{fail_count} ({100-pass_rate:.1f}%)"],
            ['Mean Raw Score:', f"{results_df['Raw Score'].str.split('/').str[0].astype(float).mean():.2f}"],
            ['Mean Percentage:', f"{results_df['Percentage'].str.rstrip('%').astype(float).mean():.1f}%"],
            ['Mean Ability (θ):', f"{results_df['Ability (θ)'].astype(float).mean():.3f}"],
            ['SD Ability (θ):', f"{results_df['Ability (θ)'].astype(float).std():.3f}"],
            ['', ''],
            ['ITEM STATISTICS', ''],
            ['Total Questions:', len(item_df)],
            ['Mean Difficulty (β):', f"{item_df['Difficulty (β)'].astype(float).mean():.3f}"],
            ['SD Difficulty (β):', f"{item_df['Difficulty (β)'].astype(float).std():.3f}"],
            ['Mean Score Rate:', f"{item_df['Avg Score %'].str.rstrip('%').astype(float).mean():.1f}%"],
            ['', ''],
            ['TIER DISTRIBUTION', ''],
        ]

        tier_counts = results_df['Tier'].value_counts()
        for tier_name in self.config['tier_names']:
            count = tier_counts.get(tier_name, 0)
            pct = (count / len(results_df) * 100) if len(results_df) > 0 else 0
            summary_data.append([tier_name + ':', f"{count} ({pct:.1f}%)"])

        for i, row_data in enumerate(summary_data, 1):
            ws.cell(row=i, column=1, value=row_data[0]).font = Font(bold=True)
            if len(row_data) > 1:
                ws.cell(row=i, column=2, value=row_data[1])

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25

    def _create_config_sheet(self, ws, config_summary):
        """Create configuration documentation sheet"""
        config_data = [
            ['ASSESSMENT CONFIGURATION', ''],
            ['', ''],
            ['QUESTION SETTINGS', ''],
            ['Number of Closed Questions:', config_summary['n_closed']],
            ['Number of Open Questions:', config_summary['n_open']],
            ['Total Questions:', config_summary['n_closed'] + config_summary['n_open']],
            ['', ''],
            ['SCORING SETTINGS', ''],
            ['Open Question Max Score:', config_summary['open_max_score']],
            ['Closed Question Weight:', config_summary['closed_weight']],
            ['Open Question Weight:', config_summary['open_weight']],
            ['', ''],
            ['GRADING SETTINGS', ''],
            ['Passing Percentage:', f"{config_summary['passing_percentage']}%"],
            ['', ''],
            ['GRADING SCALE (Score Percentage)', ''],
        ]

        thresholds = config_summary.get('tier_thresholds', [93, 87, 80, 72, 67, 60])
        tier_names = config_summary['tier_names']

        for i, threshold in enumerate(thresholds):
            next_threshold = thresholds[i-1] if i > 0 else 100
            config_data.append([tier_names[i] + ':', f"≥{threshold}%"])
        config_data.append([tier_names[-1] + ':', f"<{thresholds[-1]}%"])

        for i, row_data in enumerate(config_data, 1):
            cell = ws.cell(row=i, column=1, value=row_data[0])
            if row_data[0] and not row_data[0].endswith(':'):
                cell.font = Font(bold=True, size=12)
            else:
                cell.font = Font(bold=True)
            if len(row_data) > 1:
                ws.cell(row=i, column=2, value=row_data[1])

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25


user_configs = {}


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Send welcome message"""
    welcome_msg = (
        "🎓 *Welcome to Rasch Assessment Analysis Bot!*\n\n"
        "This bot analyzes student test results using the *Partial Credit Model (PCM)* "
        "- an advanced IRT model that properly handles both multiple choice and essay questions.\n\n"
        "*Features:*\n"
        "✅ Proper statistical model for open-ended questions\n"
        "✅ Customizable grading parameters\n"
        "✅ Fair tier assignment (same score = same tier)\n"
        "✅ Pass/Fail status based on your threshold\n"
        "✅ Comprehensive Excel reports\n\n"
        "*Commands:*\n"
        "/start - Show this message\n"
        "/config - Configure assessment parameters\n"
        "/help - Get detailed instructions\n"
        "/analyze - Start analysis with current settings\n\n"
        "Use /config to set up your assessment, then send your Excel file! 📊"
    )
    await update.message.reply_text(welcome_msg, parse_mode='Markdown')


async def config_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Start configuration wizard"""
    user_id = update.effective_user.id

    if user_id not in user_configs:
        user_configs[user_id] = DEFAULT_CONFIG.copy()

    config = user_configs[user_id]

    thresholds = config.get('tier_thresholds', [93, 87, 80, 72, 67, 60])

    config_msg = (
        "⚙️ *Current Configuration:*\n\n"
        f"📝 Closed Questions: {config['n_closed']}\n"
        f"📝 Open Questions: {config['n_open']}\n"
        f"🔢 Open Question Max Score: {config['open_max_score']}\n"
        f"✅ Passing Threshold: {config['passing_percentage']}%\n"
        f"⚖️ Closed/Open Weights: {config['closed_weight']}/{config['open_weight']}\n\n"
        f"*Grading Scale:*\n"
        f"A+: ≥{thresholds[0]}% | A: ≥{thresholds[1]}% | B+: ≥{thresholds[2]}%\n"
        f"B: ≥{thresholds[3]}% | C+: ≥{thresholds[4]}% | C: ≥{thresholds[5]}% | D: <{thresholds[5]}%\n\n"
        "*What would you like to configure?*"
    )

    keyboard = [
        [InlineKeyboardButton("📝 Number of Questions", callback_data='config_questions')],
        [InlineKeyboardButton("🔢 Open Question Max Score", callback_data='config_max_score')],
        [InlineKeyboardButton("✅ Passing Threshold", callback_data='config_passing')],
        [InlineKeyboardButton("⚖️ Question Weights", callback_data='config_weights')],
        [InlineKeyboardButton("📊 Grade Thresholds", callback_data='config_tiers')],
        [InlineKeyboardButton("🔄 Reset to Defaults", callback_data='config_reset')],
        [InlineKeyboardButton("✅ Done - Use Current Settings", callback_data='config_done')],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(config_msg, parse_mode='Markdown', reply_markup=reply_markup)


async def config_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle configuration button callbacks"""
    query = update.callback_query
    await query.answer()

    user_id = update.effective_user.id
    if user_id not in user_configs:
        user_configs[user_id] = DEFAULT_CONFIG.copy()

    data = query.data

    if data == 'config_questions':
        await query.edit_message_text(
            "📝 *Configure Number of Questions*\n\n"
            "Please send the number of closed and open questions in this format:\n"
            "`closed,open`\n\n"
            "Example: `35,10` for 35 closed and 10 open questions\n\n"
            "Or send /config to go back.",
            parse_mode='Markdown'
        )
        context.user_data['awaiting'] = 'questions'

    elif data == 'config_max_score':
        await query.edit_message_text(
            "🔢 *Configure Open Question Max Score*\n\n"
            "What is the maximum score for each open question?\n\n"
            "Send a number (e.g., `10` for 0-10 scoring, or `100` for 0-100 scoring)\n\n"
            "Or send /config to go back.",
            parse_mode='Markdown'
        )
        context.user_data['awaiting'] = 'max_score'

    elif data == 'config_passing':
        await query.edit_message_text(
            "✅ *Configure Passing Threshold*\n\n"
            "What percentage is required to pass?\n\n"
            "Send a number (e.g., `50` for 50%, or `60` for 60%)\n\n"
            "Or send /config to go back.",
            parse_mode='Markdown'
        )
        context.user_data['awaiting'] = 'passing'

    elif data == 'config_weights':
        await query.edit_message_text(
            "⚖️ *Configure Question Weights*\n\n"
            "Set the relative weights for closed and open questions.\n\n"
            "Send in format: `closed_weight,open_weight`\n"
            "Example: `1,2` means open questions are worth 2x closed questions\n"
            "Example: `1,1` means equal weight\n\n"
            "Or send /config to go back.",
            parse_mode='Markdown'
        )
        context.user_data['awaiting'] = 'weights'

    elif data == 'config_tiers':
        config = user_configs[user_id]
        current = ','.join(map(str, config.get('tier_thresholds', [93, 87, 80, 72, 67, 60])))
        await query.edit_message_text(
            "📊 *Configure Grade Thresholds*\n\n"
            f"Current thresholds: `{current}`\n\n"
            "These are minimum percentage scores for each grade.\n"
            "Send 6 numbers separated by commas (highest to lowest).\n\n"
            "Example: `93,87,80,72,67,60`\n"
            "- A+: ≥93%\n"
            "- A: ≥87%\n"
            "- B+: ≥80%\n"
            "- B: ≥72%\n"
            "- C+: ≥67%\n"
            "- C: ≥60%\n"
            "- D: <60%\n\n"
            "Or send /config to go back.",
            parse_mode='Markdown'
        )
        context.user_data['awaiting'] = 'tiers'

    elif data == 'config_reset':
        user_configs[user_id] = DEFAULT_CONFIG.copy()
        await query.edit_message_text(
            "🔄 Configuration reset to defaults!\n\n"
            "Send /config to see current settings.",
            parse_mode='Markdown'
        )
        context.user_data['awaiting'] = None

    elif data == 'config_done':
        await query.edit_message_text(
            "✅ *Configuration saved!*\n\n"
            "Now send your Excel file to analyze.\n\n"
            "Or use /config to modify settings.",
            parse_mode='Markdown'
        )
        context.user_data['awaiting'] = None


async def handle_config_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle text input for configuration"""
    awaiting = context.user_data.get('awaiting')
    if not awaiting:
        return False

    user_id = update.effective_user.id
    if user_id not in user_configs:
        user_configs[user_id] = DEFAULT_CONFIG.copy()

    text = update.message.text.strip()

    try:
        if awaiting == 'questions':
            parts = text.split(',')
            if len(parts) != 2:
                raise ValueError("Please provide exactly 2 numbers")
            n_closed = int(parts[0].strip())
            n_open = int(parts[1].strip())
            if n_closed < 0 or n_open < 0:
                raise ValueError("Numbers must be positive")
            user_configs[user_id]['n_closed'] = n_closed
            user_configs[user_id]['n_open'] = n_open
            await update.message.reply_text(
                f"✅ Updated: {n_closed} closed questions, {n_open} open questions\n\n"
                "Send /config to configure more settings or send your Excel file.",
                parse_mode='Markdown'
            )

        elif awaiting == 'max_score':
            max_score = int(text)
            if max_score < 1:
                raise ValueError("Max score must be at least 1")
            user_configs[user_id]['open_max_score'] = max_score
            await update.message.reply_text(
                f"✅ Updated: Open question max score = {max_score}\n\n"
                "Send /config to configure more settings or send your Excel file.",
                parse_mode='Markdown'
            )

        elif awaiting == 'passing':
            passing = float(text)
            if passing < 0 or passing > 100:
                raise ValueError("Percentage must be between 0 and 100")
            user_configs[user_id]['passing_percentage'] = passing
            await update.message.reply_text(
                f"✅ Updated: Passing threshold = {passing}%\n\n"
                "Send /config to configure more settings or send your Excel file.",
                parse_mode='Markdown'
            )

        elif awaiting == 'weights':
            parts = text.split(',')
            if len(parts) != 2:
                raise ValueError("Please provide exactly 2 numbers")
            closed_weight = float(parts[0].strip())
            open_weight = float(parts[1].strip())
            if closed_weight <= 0 or open_weight <= 0:
                raise ValueError("Weights must be positive")
            user_configs[user_id]['closed_weight'] = closed_weight
            user_configs[user_id]['open_weight'] = open_weight
            await update.message.reply_text(
                f"✅ Updated: Closed weight = {closed_weight}, Open weight = {open_weight}\n\n"
                "Send /config to configure more settings or send your Excel file.",
                parse_mode='Markdown'
            )

        elif awaiting == 'tiers':
            parts = text.split(',')
            if len(parts) != 6:
                raise ValueError("Please provide exactly 6 numbers")
            thresholds = [float(p.strip()) for p in parts]
            if not all(0 <= t <= 100 for t in thresholds):
                raise ValueError("All values must be between 0 and 100")
            if thresholds != sorted(thresholds, reverse=True):
                raise ValueError("Values must be in descending order (highest first)")
            user_configs[user_id]['tier_thresholds'] = thresholds
            await update.message.reply_text(
                f"✅ Updated grade thresholds: A+≥{thresholds[0]}%, A≥{thresholds[1]}%, B+≥{thresholds[2]}%, B≥{thresholds[3]}%, C+≥{thresholds[4]}%, C≥{thresholds[5]}%\n\n"
                "Send /config to configure more settings or send your Excel file.",
                parse_mode='Markdown'
            )

        context.user_data['awaiting'] = None
        return True

    except ValueError as e:
        await update.message.reply_text(
            f"❌ Invalid input: {str(e)}\n\nPlease try again or send /config to go back.",
            parse_mode='Markdown'
        )
        return True

    return False


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Send help information"""
    help_msg = (
        "📋 *Excel File Format:*\n\n"
        "*Structure:*\n"
        "```\n"
        "Student ID | Q1  | Q2  | ... | Qn\n"
        "ANSWERS    | A   | B   |     | 8\n"
        "Student1   | A   | C   |     | 7\n"
        "Student2   | B   | B   |     | 9\n"
        "```\n\n"
        "*Important Notes:*\n"
        "- First row MUST be correct answers\n"
        "- Closed questions: Letter answers (A/B/C/D/E)\n"
        "- Open questions: Integer scores (0 to max score)\n"
        "- Fractional scores are rounded to nearest integer\n"
        "- Student ID in first column\n\n"
        "*Configuration (/config):*\n"
        "- Set number of closed/open questions\n"
        "- Set max score for open questions\n"
        "- Set passing threshold percentage\n"
        "- Adjust question weights\n"
        "- Customize tier boundaries\n\n"
        "*Output Includes:*\n"
        "✅ Student rankings (same score = same rank)\n"
        "✅ Tier classifications (same % = same tier)\n"
        "✅ Pass/Fail status\n"
        "✅ Rasch ability estimates (θ)\n"
        "✅ Item difficulty parameters (β)\n"
        "✅ Statistical summaries\n"
        "✅ Configuration documentation\n\n"
        "Configure with /config, then send your Excel file! 🚀"
    )
    await update.message.reply_text(help_msg, parse_mode='Markdown')


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle uploaded Excel file"""
    try:
        user_id = update.effective_user.id
        config = user_configs.get(user_id, DEFAULT_CONFIG.copy())

        processing_msg = await update.message.reply_text(
            "⏳ Processing your file...\n"
            "Running Partial Credit Model calibration.\n"
            "This may take a moment for large datasets."
        )

        file = await update.message.document.get_file()
        file_bytes = await file.download_as_bytearray()

        processor = AssessmentProcessor(config)
        student_data = processor.parse_excel(bytes(file_bytes))

        scored_data = processor.score_responses(student_data)

        response_matrix = np.array([d['responses'] for d in scored_data])
        max_scores = processor.get_max_scores()

        if len(scored_data) < 10:
            await processing_msg.edit_text(
                "⚠️ Warning: Less than 10 students detected.\n"
                "Statistical estimates may be less stable.\n"
                "Recommended: 30+ students for reliable results.\n\n"
                "Proceeding with analysis..."
            )

        abilities, difficulties, step_difficulties = processor.pcm.calibrate(response_matrix, max_scores)

        percentages = np.array([d['percentage'] for d in scored_data])
        tiers = processor.create_tiers(abilities, percentages)

        results_df, item_df = processor.generate_report(
            scored_data, abilities, difficulties, tiers
        )

        excel_output = processor.create_excel_output(results_df, item_df, config)

        warnings = []
        if not processor.pcm.convergence_success:
            warnings.append("Model may not have fully converged - results are approximate")
        if processor.has_fractional_warning:
            warnings.append("Some fractional scores were rounded to nearest integer for statistical analysis")

        convergence_note = ""
        if warnings:
            convergence_note = "\n⚠️ " + "; ".join(warnings)

        await processing_msg.edit_text("✅ Analysis complete! Sending results...")

        pass_count = sum(1 for _, row in results_df.iterrows() if row['Status'] == 'PASS')
        pass_rate = (pass_count / len(results_df) * 100) if len(results_df) > 0 else 0

        await update.message.reply_document(
            document=excel_output,
            filename=f"PCM_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            caption=(
                f"📊 *Analysis Complete!*\n\n"
                f"👥 Students Analyzed: {len(scored_data)}\n"
                f"📝 Questions: {processor.total_questions} ({config['n_closed']} closed, {config['n_open']} open)\n"
                f"✅ Passed: {pass_count} ({pass_rate:.1f}%)\n"
                f"❌ Failed: {len(scored_data) - pass_count} ({100-pass_rate:.1f}%)\n"
                f"📈 Mean Ability: {np.mean(abilities):.3f}\n"
                f"📉 Mean Difficulty: {np.mean(difficulties):.3f}"
                f"{convergence_note}\n\n"
                f"Check the Excel file for detailed results!"
            ),
            parse_mode='Markdown'
        )

        top_5 = results_df.head(5)
        summary = "*🏆 Top 5 Students:*\n\n"
        for _, row in top_5.iterrows():
            summary += f"{row['Rank']}. {row['Student ID']} - {row['Tier']}\n   Score: {row['Percentage']} | Status: {row['Status']}\n\n"

        await update.message.reply_text(summary, parse_mode='Markdown')

    except Exception as e:
        logger.error(f"Error processing file: {e}", exc_info=True)
        await update.message.reply_text(
            f"❌ *Error processing file:*\n\n"
            f"```\n{str(e)}\n```\n\n"
            f"Please check:\n"
            f"- Excel format matches /help guidelines\n"
            f"- Number of questions matches /config settings\n"
            f"- Open question scores are within max score range",
            parse_mode='Markdown'
        )


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle text messages"""
    if await handle_config_input(update, context):
        return

    await update.message.reply_text(
        "Please send an Excel file (.xlsx) with your assessment data.\n"
        "Use /config to set up parameters first.\n"
        "Use /help to see the required format."
    )


async def main():
    """Правильный запуск для python-telegram-bot v22+"""
    application = (
        Application.builder()
        .token(BOT_TOKEN)
        .connect_timeout(30)
        .read_timeout(30)
        .pool_timeout(60)
        .build()
    )

    # Добавляешь все свои хендлеры
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("config", config_command))
    application.add_handler(CallbackQueryHandler(config_callback, pattern='^config_'))
    application.add_handler(MessageHandler(filters.Document.FileExtension("xlsx"), handle_document))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    print("Bot is starting...")

    # Эти 3 строки обязательны в v22!
    await application.initialize()
    await application.start()
    await application.updater.start_polling(drop_pending_updates=True)

    print("Bot is running! Press Ctrl+C to stop")

    # Держим бота живым
    try:
        await asyncio.Event().wait()
    except (KeyboardInterrupt, SystemExit):
        pass
    finally:
        await application.updater.stop()
        await application.stop()
        await application.shutdown()


if __name__ == '__main__':
    import asyncio
    asyncio.run(main())
